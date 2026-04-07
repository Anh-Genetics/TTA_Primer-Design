"""Module 11 — ReportGenerator: tạo báo cáo đầu ra đa định dạng.

Output files:
    1. results_summary.csv   — Tổng hợp tất cả primer pairs
    2. results_detailed.xlsx — Excel: Summary, Primers, Probes, BLAST_hits, SNPs
    3. primers.fasta         — Sequences để đặt hàng tổng hợp
    4. results.json          — Full data cho downstream tools
    5. report.html           — Báo cáo trực quan (Jinja2 template)

Visualizations (theo kế hoạch, Sprint 4):
    - Amplicon position map trên transcript/gene
    - Tm distribution chart
    - GC% scatter plot
    - Specificity score bar chart
"""

from __future__ import annotations

import csv
import json
import logging
from dataclasses import asdict
from datetime import UTC, datetime
from pathlib import Path

from tta_primer_design.config import AppConfig
from tta_primer_design.models import DesignResult

logger = logging.getLogger("tta_primer_design.modules.report_generator")

_SUPPORTED_FORMATS = {"csv", "xlsx", "json", "fasta", "html"}

# Các cột tóm tắt cho CSV / sheet Summary
_SUMMARY_FIELDS = [
    "target_id",
    "pair_id",
    "rank",
    "score",
    "amplicon_size",
    "left_seq",
    "left_tm",
    "left_gc",
    "right_seq",
    "right_tm",
    "right_gc",
    "probe_seq",
    "probe_tm",
    "probe_gc",
    "pair_penalty",
    "snp_flags",
    "status",
]


def _iter_pair_rows(results: list[DesignResult]) -> list[dict]:
    """Duyệt qua tất cả DesignResult → list các dict hàng."""
    rows = []
    for result in results:
        for rank, pair in enumerate(result.primer_pairs, start=1):
            probe = pair.probe
            row = {
                "target_id": result.target.target_id,
                "pair_id": pair.pair_id,
                "rank": rank,
                "score": pair.score,
                "amplicon_size": pair.amplicon_size,
                "left_seq": pair.left_primer.sequence,
                "left_tm": pair.left_primer.tm,
                "left_gc": pair.left_primer.gc_percent,
                "right_seq": pair.right_primer.sequence,
                "right_tm": pair.right_primer.tm,
                "right_gc": pair.right_primer.gc_percent,
                "probe_seq": probe.sequence if probe else "",
                "probe_tm": probe.tm if probe else "",
                "probe_gc": probe.gc_percent if probe else "",
                "pair_penalty": pair.pair_penalty,
                "snp_flags": "; ".join(pair.snp_flags),
                "status": result.status,
            }
            rows.append(row)
    return rows


class ReportGenerator:
    """Tạo báo cáo kết quả thiết kế primer.

    Args:
        config: AppConfig đã load.

    Example::

        reporter = ReportGenerator(config)
        reporter.generate(results, output_dir="results/run_001/")
    """

    def __init__(self, config: AppConfig) -> None:
        self.config = config

    def generate(
        self,
        results: list[DesignResult],
        output_dir: str | Path,
    ) -> list[Path]:
        """Tạo tất cả báo cáo theo config.output.formats.

        Args:
            results: Danh sách DesignResult.
            output_dir: Thư mục lưu output.

        Returns:
            Danh sách đường dẫn file đã tạo.
        """
        out_path = Path(output_dir)
        out_path.mkdir(parents=True, exist_ok=True)

        generated: list[Path] = []
        for fmt in self.config.output.formats:
            fmt = fmt.lower()
            if fmt not in _SUPPORTED_FORMATS:
                logger.warning("Unsupported output format: %s — skipping", fmt)
                continue

            try:
                if fmt == "csv":
                    path = self.generate_csv(results, out_path / "results_summary.csv")
                elif fmt == "json":
                    path = self.generate_json(results, out_path / "results.json")
                elif fmt == "xlsx":
                    path = self.generate_excel(results, out_path / "results_detailed.xlsx")
                elif fmt == "fasta":
                    path = self.generate_fasta(results, out_path / "primers.fasta")
                elif fmt == "html":
                    path = self.generate_html(results, out_path / "report.html")
                else:
                    continue

                generated.append(path)
                logger.info("Generated report: %s", path.name)

            except NotImplementedError:
                logger.debug("Report format '%s' not yet implemented", fmt)
            except Exception as exc:  # noqa: BLE001
                logger.error("Failed to generate '%s' report: %s", fmt, exc)

        return generated

    def generate_csv(self, results: list[DesignResult], output_path: Path) -> Path:
        """Tạo CSV tổng hợp.

        Args:
            results: Danh sách DesignResult.
            output_path: Đường dẫn file output.

        Returns:
            Đường dẫn file đã tạo.
        """
        rows = _iter_pair_rows(results)
        with output_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=_SUMMARY_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
        return output_path

    def generate_excel(self, results: list[DesignResult], output_path: Path) -> Path:
        """Tạo Excel với nhiều sheet.

        Sheets:
            - Summary: tóm tắt tất cả primer pairs
            - Primers: thông tin chi tiết oligo
            - Probes: thông tin probe (nếu có)
            - SNPs: danh sách SNP flags
            - BLAST_hits: thông tin specificity (nếu có)

        Args:
            results: Danh sách DesignResult.
            output_path: Đường dẫn file .xlsx.

        Returns:
            Đường dẫn file đã tạo.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise ImportError("openpyxl là bắt buộc cho generate_excel()") from exc

        wb = openpyxl.Workbook()

        # ---- Sheet: Summary ----
        ws_summary = wb.active
        ws_summary.title = "Summary"
        _write_ws_rows(ws_summary, _SUMMARY_FIELDS, _iter_pair_rows(results))
        _style_header_row(ws_summary, Font, PatternFill, Alignment)

        # ---- Sheet: Primers ----
        ws_primers = wb.create_sheet("Primers")
        primer_fields = [
            "target_id",
            "pair_id",
            "oligo",
            "sequence",
            "start",
            "length",
            "tm",
            "gc_percent",
            "self_any_th",
            "self_end_th",
            "hairpin_th",
            "end_stability",
            "penalty",
        ]
        primer_rows: list[dict] = []
        for result in results:
            for pair in result.primer_pairs:
                for label, oligo in [
                    ("LEFT", pair.left_primer),
                    ("RIGHT", pair.right_primer),
                ]:
                    primer_rows.append(
                        {
                            "target_id": result.target.target_id,
                            "pair_id": pair.pair_id,
                            "oligo": label,
                            "sequence": oligo.sequence,
                            "start": oligo.start,
                            "length": oligo.length,
                            "tm": oligo.tm,
                            "gc_percent": oligo.gc_percent,
                            "self_any_th": oligo.self_any_th,
                            "self_end_th": oligo.self_end_th,
                            "hairpin_th": oligo.hairpin_th,
                            "end_stability": oligo.end_stability,
                            "penalty": oligo.penalty,
                        }
                    )
        _write_ws_rows(ws_primers, primer_fields, primer_rows)
        _style_header_row(ws_primers, Font, PatternFill, Alignment)

        # ---- Sheet: Probes ----
        ws_probes = wb.create_sheet("Probes")
        probe_fields = [
            "target_id",
            "pair_id",
            "sequence",
            "start",
            "length",
            "tm",
            "gc_percent",
        ]
        probe_rows: list[dict] = []
        for result in results:
            for pair in result.primer_pairs:
                if pair.probe is not None:
                    probe_rows.append(
                        {
                            "target_id": result.target.target_id,
                            "pair_id": pair.pair_id,
                            "sequence": pair.probe.sequence,
                            "start": pair.probe.start,
                            "length": pair.probe.length,
                            "tm": pair.probe.tm,
                            "gc_percent": pair.probe.gc_percent,
                        }
                    )
        _write_ws_rows(ws_probes, probe_fields, probe_rows)
        _style_header_row(ws_probes, Font, PatternFill, Alignment)

        # ---- Sheet: SNPs ----
        ws_snps = wb.create_sheet("SNPs")
        snp_fields = ["target_id", "pair_id", "flag"]
        snp_rows: list[dict] = []
        for result in results:
            for pair in result.primer_pairs:
                for flag in pair.snp_flags:
                    snp_rows.append(
                        {
                            "target_id": result.target.target_id,
                            "pair_id": pair.pair_id,
                            "flag": flag,
                        }
                    )
        _write_ws_rows(ws_snps, snp_fields, snp_rows)
        _style_header_row(ws_snps, Font, PatternFill, Alignment)

        # ---- Sheet: BLAST_hits ----
        ws_blast = wb.create_sheet("BLAST_hits")
        blast_fields = ["target_id", "pair_id", "is_specific", "specificity_score", "off_targets"]
        blast_rows: list[dict] = []
        for result in results:
            for pair in result.primer_pairs:
                if pair.specificity_result is not None:
                    sr = pair.specificity_result
                    blast_rows.append(
                        {
                            "target_id": result.target.target_id,
                            "pair_id": pair.pair_id,
                            "is_specific": getattr(sr, "is_specific", ""),
                            "specificity_score": getattr(sr, "specificity_score", ""),
                            "off_targets": len(getattr(sr, "off_target_amplicons", [])),
                        }
                    )
        _write_ws_rows(ws_blast, blast_fields, blast_rows)
        _style_header_row(ws_blast, Font, PatternFill, Alignment)

        wb.save(output_path)
        return output_path

    def generate_json(self, results: list[DesignResult], output_path: Path) -> Path:
        """Tạo JSON đầy đủ.

        Args:
            results: Danh sách DesignResult.
            output_path: Đường dẫn file .json.

        Returns:
            Đường dẫn file đã tạo.
        """
        serializable = []
        for r in results:
            try:
                serializable.append(asdict(r))
            except Exception as exc:  # noqa: BLE001
                logger.warning("Could not serialize result for %s: %s", r.target.target_id, exc)
                serializable.append({"target_id": r.target.target_id, "status": r.status})

        with output_path.open("w", encoding="utf-8") as fh:
            json.dump(serializable, fh, indent=2, ensure_ascii=False, default=str)

        return output_path

    def generate_fasta(self, results: list[DesignResult], output_path: Path) -> Path:
        """Tạo FASTA file chứa tất cả oligo sequences.

        Header format: >{target_id}_{pair_id}_{oligo_type}

        Args:
            results: Danh sách DesignResult.
            output_path: Đường dẫn file .fasta.

        Returns:
            Đường dẫn file đã tạo.
        """
        with output_path.open("w", encoding="utf-8") as fh:
            for result in results:
                for pair in result.primer_pairs:
                    tid = result.target.target_id
                    pid = pair.pair_id

                    fh.write(f">{tid}_{pid}_LEFT\n{pair.left_primer.sequence}\n")
                    fh.write(f">{tid}_{pid}_RIGHT\n{pair.right_primer.sequence}\n")
                    if pair.probe is not None:
                        fh.write(f">{tid}_{pid}_PROBE\n{pair.probe.sequence}\n")

        return output_path

    def generate_html(self, results: list[DesignResult], output_path: Path) -> Path:
        """Tạo báo cáo HTML với Jinja2.

        Args:
            results: Danh sách DesignResult.
            output_path: Đường dẫn file .html.

        Returns:
            Đường dẫn file đã tạo.
        """
        try:
            from jinja2 import Environment
        except ImportError as exc:
            raise ImportError("jinja2 là bắt buộc cho generate_html()") from exc

        env = Environment(autoescape=True)
        template = env.from_string(_HTML_TEMPLATE)

        total_targets = len(results)
        total_pairs = sum(len(r.primer_pairs) for r in results)
        success_count = sum(1 for r in results if r.status == "success")
        generated_at = datetime.now(tz=UTC).strftime("%Y-%m-%d %H:%M UTC")

        html_content = template.render(
            results=results,
            total_targets=total_targets,
            total_pairs=total_pairs,
            success_count=success_count,
            generated_at=generated_at,
            pipeline_name=self.config.pipeline.name,
            pipeline_version=self.config.pipeline.version,
            pipeline_mode=self.config.pipeline.mode,
        )

        with output_path.open("w", encoding="utf-8") as fh:
            fh.write(html_content)

        return output_path


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _write_ws_rows(ws, fields: list[str], rows: list[dict]) -> None:  # type: ignore[no-untyped-def]
    """Ghi header + rows vào worksheet."""
    ws.append(fields)
    for row in rows:
        ws.append([row.get(f, "") for f in fields])


def _style_header_row(ws, Font, PatternFill, Alignment) -> None:  # type: ignore[no-untyped-def]
    """Áp dụng định dạng cho hàng header."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="366092")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Jinja2 HTML template (inline — không cần file template bên ngoài)
# ---------------------------------------------------------------------------

_HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="vi">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{{ pipeline_name }} — Primer Design Report</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 20px; color: #333; }
    h1 { color: #366092; }
    h2 { color: #4472C4; border-bottom: 2px solid #4472C4; padding-bottom: 4px; }
    .summary-box {
      background: #EBF0FA; border: 1px solid #4472C4;
      border-radius: 6px; padding: 12px 20px; margin-bottom: 20px;
      display: inline-block;
    }
    .summary-box span { margin-right: 24px; font-weight: bold; }
    table { border-collapse: collapse; width: 100%; margin-bottom: 24px; font-size: 13px; }
    th { background: #366092; color: #fff; padding: 6px 10px; text-align: left; }
    td { padding: 5px 10px; border-bottom: 1px solid #ddd; }
    tr:nth-child(even) { background: #f5f7fc; }
    .badge-success { background:#27AE60; color:#fff; border-radius:3px; padding:2px 7px; }
    .badge-failed  { background:#E74C3C; color:#fff; border-radius:3px; padding:2px 7px; }
    .badge-no-primers { background:#F39C12; color:#fff; border-radius:3px; padding:2px 7px; }
    .score-high { color: #27AE60; font-weight: bold; }
    .score-mid  { color: #F39C12; font-weight: bold; }
    .score-low  { color: #E74C3C; font-weight: bold; }
    footer { color: #888; font-size: 12px; margin-top: 30px; }
  </style>
</head>
<body>
<h1>{{ pipeline_name }} v{{ pipeline_version }} — Primer Design Report</h1>
<div class="summary-box">
  <span>Mode: {{ pipeline_mode }}</span>
  <span>Targets: {{ total_targets }}</span>
  <span>Succeeded: {{ success_count }}/{{ total_targets }}</span>
  <span>Primer pairs: {{ total_pairs }}</span>
  <span>Generated: {{ generated_at }}</span>
</div>

{% for result in results %}
<h2>Target: {{ result.target.target_id }}
  {% if result.status == "success" %}<span class="badge-success">success</span>
  {% elif result.status == "failed" %}<span class="badge-failed">failed</span>
  {% else %}<span class="badge-no-primers">{{ result.status }}</span>{% endif %}
</h2>

{% if result.error %}
<p style="color:#E74C3C"><strong>Error:</strong> {{ result.error }}</p>
{% endif %}

{% if result.primer_pairs %}
<table>
  <thead>
    <tr>
      <th>#</th><th>Pair ID</th><th>Score</th><th>Amplicon (bp)</th>
      <th>LEFT sequence</th><th>Tm L</th><th>GC% L</th>
      <th>RIGHT sequence</th><th>Tm R</th><th>GC% R</th>
      <th>Probe</th><th>SNP flags</th>
    </tr>
  </thead>
  <tbody>
  {% for pair in result.primer_pairs %}
    <tr>
      <td>{{ loop.index }}</td>
      <td>{{ pair.pair_id }}</td>
      <td>
        {% if pair.score >= 70 %}<span class="score-high">{{ "%.1f"|format(pair.score) }}</span>
        {% elif pair.score >= 40 %}<span class="score-mid">{{ "%.1f"|format(pair.score) }}</span>
        {% else %}<span class="score-low">{{ "%.1f"|format(pair.score) }}</span>{% endif %}
      </td>
      <td>{{ pair.amplicon_size if pair.amplicon_size else "—" }}</td>
      <td><code>{{ pair.left_primer.sequence }}</code></td>
      <td>{{ "%.1f"|format(pair.left_primer.tm) if pair.left_primer.tm else "—" }}</td>
      <td>{{ "%.1f"|format(pair.left_primer.gc_percent) if pair.left_primer.gc_percent else "—" }}</td>
      <td><code>{{ pair.right_primer.sequence }}</code></td>
      <td>{{ "%.1f"|format(pair.right_primer.tm) if pair.right_primer.tm else "—" }}</td>
      <td>{{ "%.1f"|format(pair.right_primer.gc_percent) if pair.right_primer.gc_percent else "—" }}</td>
      <td>{% if pair.probe %}<code>{{ pair.probe.sequence }}</code>{% else %}—{% endif %}</td>
      <td>{{ pair.snp_flags | join("; ") if pair.snp_flags else "—" }}</td>
    </tr>
  {% endfor %}
  </tbody>
</table>
{% else %}
<p><em>Không có primer pairs nào.</em></p>
{% endif %}
{% endfor %}

<footer>
  <p>Generated by {{ pipeline_name }} v{{ pipeline_version }} &mdash; {{ generated_at }}</p>
</footer>
</body>
</html>
"""
