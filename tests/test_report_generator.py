"""Tests cho ReportGenerator — CSV, Excel, FASTA, JSON, HTML."""

from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest

from tta_primer_design.config import AppConfig, OutputConfig, load_config
from tta_primer_design.models import DesignResult, DesignTarget, Oligo, PrimerPair
from tta_primer_design.modules.report_generator import ReportGenerator

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def config() -> AppConfig:
    cfg = load_config(None)
    cfg.output = OutputConfig(formats=["csv", "json", "xlsx", "fasta", "html"])
    return cfg


@pytest.fixture
def reporter(config: AppConfig) -> ReportGenerator:
    return ReportGenerator(config)


def _make_target(tid: str = "ACTB") -> DesignTarget:
    return DesignTarget(target_id=tid, accession="NM_001101")


def _make_oligo(seq: str = "ATCGATCGATCGATCGATCG", tm: float = 60.0, gc: float = 50.0) -> Oligo:
    return Oligo(sequence=seq, tm=tm, gc_percent=gc)


def _make_pair(
    pair_id: str = "pair_0",
    amplicon_size: int = 120,
    snp_flags: list[str] | None = None,
    with_probe: bool = False,
    score: float = 75.0,
) -> PrimerPair:
    probe = _make_oligo(seq="TTTTGGGGCCCCAAAATTTT", tm=67.0) if with_probe else None
    return PrimerPair(
        pair_id=pair_id,
        left_primer=_make_oligo(),
        right_primer=_make_oligo(seq="GCTAGCTAGCTAGCTAGCTA"),
        amplicon_size=amplicon_size,
        snp_flags=snp_flags or [],
        score=score,
        probe=probe,
    )


def _make_result(
    tid: str = "ACTB",
    n_pairs: int = 2,
    status: str = "success",
    with_probe: bool = False,
) -> DesignResult:
    target = _make_target(tid)
    pairs = [_make_pair(f"pair_{i}", with_probe=with_probe) for i in range(n_pairs)]
    return DesignResult(target=target, primer_pairs=pairs, status=status)


# ---------------------------------------------------------------------------
# Tests: generate_csv
# ---------------------------------------------------------------------------


class TestGenerateCSV:
    def test_creates_file(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_csv(results, tmp_path / "out.csv")
        assert path.exists()

    def test_returns_path(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_csv(results, tmp_path / "out.csv")
        assert isinstance(path, Path)

    def test_header_row(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_csv(results, tmp_path / "out.csv")
        with path.open() as fh:
            reader = csv.DictReader(fh)
            assert "target_id" in reader.fieldnames  # type: ignore[operator]
            assert "pair_id" in reader.fieldnames  # type: ignore[operator]
            assert "score" in reader.fieldnames  # type: ignore[operator]

    def test_row_count(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result(n_pairs=3)]
        path = reporter.generate_csv(results, tmp_path / "out.csv")
        with path.open() as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
        assert len(rows) == 3

    def test_multiple_targets(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result("A", n_pairs=2), _make_result("B", n_pairs=1)]
        path = reporter.generate_csv(results, tmp_path / "out.csv")
        with path.open() as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
        assert len(rows) == 3

    def test_target_id_in_rows(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result("GAPDH", n_pairs=1)]
        path = reporter.generate_csv(results, tmp_path / "out.csv")
        with path.open() as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
        assert rows[0]["target_id"] == "GAPDH"

    def test_empty_results(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        path = reporter.generate_csv([], tmp_path / "out.csv")
        assert path.exists()
        with path.open() as fh:
            reader = csv.DictReader(fh)
            rows = list(reader)
        assert rows == []

    def test_snp_flags_in_csv(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        pair = _make_pair(snp_flags=["WARNING:mid", "WARNING:dg"])
        result = DesignResult(target=_make_target(), primer_pairs=[pair], status="success")
        path = reporter.generate_csv([result], tmp_path / "out.csv")
        with path.open() as fh:
            reader = csv.DictReader(fh)
            row = next(reader)
        assert "WARNING" in row["snp_flags"]

    def test_probe_seq_in_csv(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        result = _make_result(with_probe=True, n_pairs=1)
        path = reporter.generate_csv([result], tmp_path / "out.csv")
        with path.open() as fh:
            reader = csv.DictReader(fh)
            row = next(reader)
        assert len(row["probe_seq"]) > 0

    def test_no_probe_empty_field(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        result = _make_result(with_probe=False, n_pairs=1)
        path = reporter.generate_csv([result], tmp_path / "out.csv")
        with path.open() as fh:
            reader = csv.DictReader(fh)
            row = next(reader)
        assert row["probe_seq"] == ""


# ---------------------------------------------------------------------------
# Tests: generate_fasta
# ---------------------------------------------------------------------------


class TestGenerateFASTA:
    def test_creates_file(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_fasta(results, tmp_path / "primers.fasta")
        assert path.exists()

    def test_contains_fasta_headers(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result("ACTB", n_pairs=1)]
        path = reporter.generate_fasta(results, tmp_path / "primers.fasta")
        content = path.read_text()
        assert ">ACTB_pair_0_LEFT" in content
        assert ">ACTB_pair_0_RIGHT" in content

    def test_sequences_present(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result(n_pairs=1)]
        path = reporter.generate_fasta(results, tmp_path / "primers.fasta")
        content = path.read_text()
        assert "ATCGATCGATCGATCGATCG" in content

    def test_probe_in_fasta(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result(with_probe=True, n_pairs=1)]
        path = reporter.generate_fasta(results, tmp_path / "primers.fasta")
        content = path.read_text()
        assert "_PROBE" in content

    def test_no_probe_no_probe_header(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result(with_probe=False, n_pairs=1)]
        path = reporter.generate_fasta(results, tmp_path / "primers.fasta")
        content = path.read_text()
        assert "_PROBE" not in content

    def test_empty_results(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        path = reporter.generate_fasta([], tmp_path / "primers.fasta")
        assert path.exists()
        assert path.read_text() == ""

    def test_multiple_pairs(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result(n_pairs=3)]
        path = reporter.generate_fasta(results, tmp_path / "primers.fasta")
        content = path.read_text()
        headers = [line for line in content.splitlines() if line.startswith(">")]
        assert len(headers) == 6  # 3 pairs × 2 oligos (LEFT + RIGHT)


# ---------------------------------------------------------------------------
# Tests: generate_json
# ---------------------------------------------------------------------------


class TestGenerateJSON:
    def test_creates_file(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_json(results, tmp_path / "results.json")
        assert path.exists()

    def test_valid_json(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_json(results, tmp_path / "results.json")
        data = json.loads(path.read_text())
        assert isinstance(data, list)

    def test_target_id_in_json(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result("TP53")]
        path = reporter.generate_json(results, tmp_path / "results.json")
        data = json.loads(path.read_text())
        assert data[0]["target"]["target_id"] == "TP53"

    def test_empty_results(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        path = reporter.generate_json([], tmp_path / "results.json")
        data = json.loads(path.read_text())
        assert data == []


# ---------------------------------------------------------------------------
# Tests: generate_excel
# ---------------------------------------------------------------------------


class TestGenerateExcel:
    def test_creates_file(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_excel(results, tmp_path / "results.xlsx")
        assert path.exists()

    def test_file_is_xlsx(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_excel(results, tmp_path / "results.xlsx")
        assert path.suffix == ".xlsx"

    def test_xlsx_has_summary_sheet(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        import openpyxl

        results = [_make_result()]
        path = reporter.generate_excel(results, tmp_path / "results.xlsx")
        wb = openpyxl.load_workbook(path)
        assert "Summary" in wb.sheetnames

    def test_xlsx_has_primers_sheet(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        import openpyxl

        results = [_make_result()]
        path = reporter.generate_excel(results, tmp_path / "results.xlsx")
        wb = openpyxl.load_workbook(path)
        assert "Primers" in wb.sheetnames

    def test_xlsx_has_probes_sheet(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        import openpyxl

        results = [_make_result()]
        path = reporter.generate_excel(results, tmp_path / "results.xlsx")
        wb = openpyxl.load_workbook(path)
        assert "Probes" in wb.sheetnames

    def test_xlsx_has_snps_sheet(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        import openpyxl

        results = [_make_result()]
        path = reporter.generate_excel(results, tmp_path / "results.xlsx")
        wb = openpyxl.load_workbook(path)
        assert "SNPs" in wb.sheetnames

    def test_xlsx_has_blast_sheet(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        import openpyxl

        results = [_make_result()]
        path = reporter.generate_excel(results, tmp_path / "results.xlsx")
        wb = openpyxl.load_workbook(path)
        assert "BLAST_hits" in wb.sheetnames

    def test_summary_row_count(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        import openpyxl

        results = [_make_result(n_pairs=2)]
        path = reporter.generate_excel(results, tmp_path / "results.xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb["Summary"]
        assert ws.max_row == 3  # 1 header + 2 data rows

    def test_empty_results(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        path = reporter.generate_excel([], tmp_path / "results.xlsx")
        assert path.exists()


# ---------------------------------------------------------------------------
# Tests: generate_html
# ---------------------------------------------------------------------------


class TestGenerateHTML:
    def test_creates_file(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_html(results, tmp_path / "report.html")
        assert path.exists()

    def test_valid_html_structure(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_html(results, tmp_path / "report.html")
        content = path.read_text(encoding="utf-8")
        assert "<!DOCTYPE html>" in content
        assert "<html" in content
        assert "</html>" in content

    def test_target_id_in_html(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result("GAPDH")]
        path = reporter.generate_html(results, tmp_path / "report.html")
        content = path.read_text(encoding="utf-8")
        assert "GAPDH" in content

    def test_pipeline_name_in_html(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result()]
        path = reporter.generate_html(results, tmp_path / "report.html")
        content = path.read_text(encoding="utf-8")
        assert reporter.config.pipeline.name in content

    def test_primer_sequences_in_html(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result(n_pairs=1)]
        path = reporter.generate_html(results, tmp_path / "report.html")
        content = path.read_text(encoding="utf-8")
        assert "ATCGATCGATCGATCGATCG" in content

    def test_empty_results(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        path = reporter.generate_html([], tmp_path / "report.html")
        content = path.read_text(encoding="utf-8")
        assert "<!DOCTYPE html>" in content

    def test_score_displayed(self, reporter: ReportGenerator, tmp_path: Path) -> None:
        results = [_make_result(n_pairs=1)]
        results[0].primer_pairs[0].score = 85.0
        path = reporter.generate_html(results, tmp_path / "report.html")
        content = path.read_text(encoding="utf-8")
        assert "85" in content


# ---------------------------------------------------------------------------
# Tests: generate (dispatcher)
# ---------------------------------------------------------------------------


class TestGenerate:
    def test_generate_csv_and_json(self, tmp_path: Path) -> None:
        cfg = load_config(None)
        cfg.output = OutputConfig(formats=["csv", "json"])
        reporter = ReportGenerator(cfg)
        results = [_make_result()]
        paths = reporter.generate(results, tmp_path)
        names = [p.name for p in paths]
        assert "results_summary.csv" in names
        assert "results.json" in names

    def test_generate_creates_output_dir(self, tmp_path: Path) -> None:
        cfg = load_config(None)
        cfg.output = OutputConfig(formats=["json"])
        reporter = ReportGenerator(cfg)
        out = tmp_path / "new_dir" / "subdir"
        reporter.generate([_make_result()], out)
        assert out.exists()

    def test_generate_unknown_format_skipped(self, tmp_path: Path) -> None:
        cfg = load_config(None)
        cfg.output = OutputConfig(formats=["unsupported_format", "json"])
        reporter = ReportGenerator(cfg)
        paths = reporter.generate([_make_result()], tmp_path)
        assert len(paths) == 1  # only json

    def test_generate_all_formats(self, tmp_path: Path) -> None:
        cfg = load_config(None)
        cfg.output = OutputConfig(formats=["csv", "json", "xlsx", "fasta", "html"])
        reporter = ReportGenerator(cfg)
        paths = reporter.generate([_make_result()], tmp_path)
        assert len(paths) == 5

    def test_generate_empty_results(self, tmp_path: Path) -> None:
        cfg = load_config(None)
        cfg.output = OutputConfig(formats=["csv", "json"])
        reporter = ReportGenerator(cfg)
        paths = reporter.generate([], tmp_path)
        assert len(paths) == 2  # files created even with empty results
